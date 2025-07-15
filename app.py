from flask import Flask, render_template, request, redirect, url_for, flash, session, g, jsonify, Response, make_response
import sqlite3
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from database import init_app, get_db
from datetime import datetime
import json
import io
import os
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill 
import firebase_admin
from firebase_admin import credentials, db

try:

    base_dir = os.path.dirname(os.path.abspath(__file__))
    credentials_path = os.path.join(base_dir, 'firebase-credentials.json')

    
    if not os.path.exists(credentials_path):
        raise FileNotFoundError(f"O arquivo de credenciais não foi encontrado em: {credentials_path}")

    cred = credentials.Certificate(credentials_path)
    firebase_admin.initialize_app(cred, {
        'databaseURL': 'https://desativar-site-default-rtdb.firebaseio.com/'
    })
    print("Firebase inicializado com sucesso.")
except Exception as e:
    print(f"ERRO: Não foi possível inicializar o Firebase. Verifique o arquivo 'firebase-credentials.json' e a URL. Erro: {e}")

app = Flask(__name__)

app.config['APPLICATION_ROOT'] = os.environ.get('APP_PREFIX', '')

app.secret_key = '09164Duque!Paprika'
PER_PAGE = 20  

app.jinja_env.add_extension('jinja2.ext.do')

@app.before_request
def check_for_maintenance():
    
    
    if request.endpoint and request.endpoint not in ['static', 'maintenance']:
        try:
            
            maintenance_ref = db.reference('config/is_maintenance_mode')
            is_maintenance = maintenance_ref.get()

            
            if is_maintenance:
                return render_template('maintenance.html'), 503 
        except Exception as e:
            
            
            print(f"AVISO: Não foi possível verificar o modo de manutenção no Firebase. O site continuará online. Erro: {e}")


@app.route('/maintenance')
def maintenance():
    return render_template('maintenance.html'), 503



@app.template_filter('datetime')
def format_datetime(value, fmt='%d/%m/%Y %H:%M:%S'):
    """Formata uma string de data/hora para o formato brasileiro."""
    if value is None:
        return ""
    try:
        return datetime.strptime(value, '%Y-%m-%d %H:%M:%S.%f').strftime(fmt)
    except ValueError:
        try:
            return datetime.strptime(value, '%Y-%m-%d %H:%M:%S').strftime(fmt)
        except ValueError:
            return value

@app.template_filter('prettyjson')
def pretty_json_filter(value):
    """Formata uma string JSON para uma exibição HTML legível."""
    if not value or value == 'null':
        return 'N/A'
    try:
        data = json.loads(value)
        
        html = "<ul>"
        for key, val in data.items():
            if key != 'password': 
                html += f"<li><strong>{key.replace('_', ' ').title()}:</strong> {val}</li>"
        html += "</ul>"
        return html
    except (json.JSONDecodeError, AttributeError):
        return value


init_app(app)



def is_ajax_request():
    """Verifica se a requisição é do tipo AJAX."""
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest'


def log_action(action, target_type=None, target_id=None, target_name=None, dados_antigos=None, dados_novos=None):
    db = get_db()
    dados_antigos_str = json.dumps(dados_antigos, ensure_ascii=False) if isinstance(dados_antigos, dict) else dados_antigos
    dados_novos_str = json.dumps(dados_novos, ensure_ascii=False) if isinstance(dados_novos, dict) else dados_novos
    db.execute(
        "INSERT INTO audit_log (user_id, username, action, target_type, target_id, target_name, dados_antigos, dados_novos) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (session.get('user_id'), session.get('username'), action, target_type, target_id, target_name, dados_antigos_str, dados_novos_str)
    )
    db.commit()



def login_required(f):
    """Garante que o usuário esteja logado para acessar a rota."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            if is_ajax_request():
                return jsonify({'status': 'error', 'message': 'Sessão expirada. Faça login novamente.'}), 401
            flash('Você precisa fazer login para acessar esta página.', 'danger')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function


def role_required(allowed_roles):
    """Garante que o usuário tenha uma das funções permitidas."""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not session.get('logged_in') or session.get('role') not in allowed_roles:
                if is_ajax_request():
                    return jsonify({'status': 'error', 'message': 'Você não tem permissão para esta ação.'}), 403
                log_action('acesso_negado', target_name=request.path)
                flash('Você não tem permissão para acessar esta página.', 'danger')
                return redirect(url_for('admin_dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator



@app.route('/api/cliente/<codigo_cliente>')
@login_required
def get_client_info(codigo_cliente):
    db = get_db()
    
    store_id_clause = "store_id = ?"
    params = [codigo_cliente, session.get('store_id')]
    
    if session.get('role') == 'super_admin':
        store_id = request.args.get('store_id')
        if not store_id:
             return jsonify({'status': 'error', 'message': 'Selecione uma empresa.'}), 400
        params = [codigo_cliente, store_id]

    
    client = db.execute(f"SELECT nome_cliente FROM clientes WHERE codigo_cliente = ? AND {store_id_clause}", params).fetchone()
    
    if client:
        return jsonify({'status': 'success', 'nome_cliente': client['nome_cliente']})
    else:
        return jsonify({'status': 'not_found'}), 404

@app.route('/api/marcas/<int:tipo_id>')
@login_required
def get_marcas_por_tipo(tipo_id):
    db = get_db()
    marcas = db.execute("SELECT id, nome FROM marcas WHERE tipo_id = ? ORDER BY nome", (tipo_id,)).fetchall()
    return jsonify([dict(m) for m in marcas])



@app.route('/')
def index():
    return render_template('index.html')


@app.route('/dashboard')
@login_required
def admin_dashboard():
    role_map = {
        'super_admin': 'super_admin_dashboard',
        'admin_patrimonio': 'patrimonio_dashboard',
        'admin_rh': 'rh_dashboard',
        'admin_contas_a_pagar': 'contas_a_pagar_dashboard',
        'admin_cobranca': 'cobranca_dashboard'
    }
    dashboard_route = role_map.get(session.get('role'), 'index')
    return redirect(url_for(dashboard_route))


@app.route('/search', methods=['GET'])
def search():
    page = request.args.get('page', 1, type=int)
    query = request.args.get('query', '').strip()
    search_by = request.args.get('search_by', 'codigo_cliente')
    all_results = []

    if query:
        db = get_db()
        query_term = '%' + query + '%'

        if search_by == 'codigo_cliente':
            all_results = db.execute("SELECT * FROM clientes WHERE codigo_cliente LIKE ?", (query_term,)).fetchall()
        elif search_by == 'nome_cliente':
            all_results = db.execute("SELECT * FROM clientes WHERE nome_cliente LIKE ?", (query_term,)).fetchall()
        elif search_by == 'patrimonio_especifico':
            
            
            
            items = db.execute("""
                SELECT c.*, t.nome as tipo, m.nome as marca, pi.tamanho, pi.codigo_patrimonio
                FROM clientes c
                JOIN patrimonio_items pi ON c.id = pi.cliente_id
                LEFT JOIN tipos_equipamento t ON pi.tipo_id = t.id
                LEFT JOIN marcas m ON pi.marca_id = m.id
                WHERE pi.codigo_patrimonio LIKE ?
            """, (query_term,)).fetchall()
            all_results = items
        elif search_by == 'caixa_cobranca_range':
            try:
                cliente_num = int(query)
                all_results = db.execute("SELECT * FROM cobranca_fichas_acerto WHERE ? BETWEEN range_cliente_inicio AND range_cliente_fim", (cliente_num,)).fetchall()
            except ValueError:
                flash('Para pesquisa por "Caixa Cobrança", digite um número de cliente válido.', 'danger')
        elif search_by == 'numero_caixa':
            pat_res = db.execute("SELECT c.numero_caixa, t.nome as tipo, m.nome as marca, s.name as store_name FROM clientes c JOIN patrimonio_items pi ON c.id = pi.cliente_id JOIN tipos_equipamento t ON pi.tipo_id = t.id JOIN marcas m ON pi.marca_id = m.id LEFT JOIN stores s ON c.store_id = s.id WHERE c.numero_caixa LIKE ?", (query_term,)).fetchall()
            for row in pat_res:
                all_results.append({'type': row['tipo'], 'caixa': row['numero_caixa'], 'description': row['marca'], 'store_name': row['store_name']})
            pag_res = db.execute("SELECT cap.caixa, cap.pagamento_data_inicio, cap.pagamento_data_fim, 'Pagamento' as type, s.name as store_name FROM contas_a_pagar_pagamentos cap LEFT JOIN stores s ON cap.store_id = s.id WHERE cap.caixa LIKE ?", (query_term,)).fetchall()
            for row in pag_res:
                all_results.append({'type': row['type'], 'caixa': row['caixa'], 'description': f"Período de {row['pagamento_data_inicio']} a {row['pagamento_data_fim']}", 'store_name': row['store_name']})
            div_res = db.execute("SELECT cad.numero_caixa, 'Documento Diverso' as type, s.name as store_name FROM contas_a_pagar_diversos cad LEFT JOIN stores s ON cad.store_id = s.id WHERE cad.numero_caixa LIKE ?", (query_term,)).fetchall()
            for row in div_res:
                all_results.append({'type': row['type'], 'caixa': row['numero_caixa'], 'description': 'N/A', 'store_name': row['store_name']})
            cob_res = db.execute("SELECT cfa.caixa, cfa.range_cliente_inicio, cfa.range_cliente_fim, 'Cobrança' as type, s.name as store_name FROM cobranca_fichas_acerto cfa LEFT JOIN stores s ON cfa.store_id = s.id WHERE cfa.caixa LIKE ?", (query_term,)).fetchall()
            for row in cob_res:
                all_results.append({'type': row['type'], 'caixa': row['caixa'], 'description': f"{row['range_cliente_inicio']} - {row['range_cliente_fim']}", 'store_name': row['store_name']})

    total_items = len(all_results)
    total_pages = (total_items + PER_PAGE - 1) // PER_PAGE if total_items > 0 else 1
    offset = (page - 1) * PER_PAGE
    paginated_results = all_results[offset : offset + PER_PAGE]

    return render_template('search_results.html', 
                           results=paginated_results, 
                           query=query, 
                           search_by=search_by,
                           page=page,
                           total_pages=total_pages)



@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        db = get_db()
        user = db.execute("SELECT u.*, s.name as store_name FROM users u LEFT JOIN stores s ON u.store_id = s.id WHERE u.username = ?", (username,)).fetchone()
        
        if user and check_password_hash(user['password'], password):
            session.clear()
            session['logged_in'] = True
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            session['store_id'] = user['store_id']
            session['store_name'] = user['store_name']
            log_action('fez_login')
            flash(f'Login bem-sucedido! Bem-vindo, {user["username"]}.', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            log_action('erro_no_login', target_name=username)
            flash('Usuário ou senha inválidos.', 'danger')
    return render_template('admin_login.html')


@app.route('/admin/logout')
def admin_logout():
    
    if 'username' in session:
        
        log_action('Saiu')
    
    
    session.clear()
    
    
    flash('Você foi desconectado com segurança.', 'info')
    
    
    response = make_response(redirect(url_for('index')))
    
    
    
    
    
    response.set_cookie('session', '', expires=0)
    
    
    return response


@app.route('/profile/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']

        db = get_db()
        user = db.execute("SELECT * FROM users WHERE id = ?", (session['user_id'],)).fetchone()

        if not user or not check_password_hash(user['password'], current_password):
            flash('Sua senha atual está incorreta.', 'danger')
        elif len(new_password) < 6:
            flash('A nova senha deve ter pelo menos 6 caracteres.', 'warning')
        elif new_password != confirm_password:
            flash('A nova senha e a confirmação não correspondem.', 'danger')
        else:
            hashed_password = generate_password_hash(new_password)
            db.execute("UPDATE users SET password = ? WHERE id = ?", (hashed_password, session['user_id']))
            db.commit()
            log_action('trocou_sua_senha')
            flash('Sua senha foi alterada com sucesso!', 'success')
            return redirect(url_for('admin_dashboard'))

    return render_template('profile/change_password.html')



@app.route('/super_admin/dashboard')
@login_required
@role_required(['super_admin'])
def super_admin_dashboard():
    db = get_db()
    
    
    total_users = db.execute("SELECT COUNT(id) FROM users").fetchone()[0]
    total_stores = db.execute("SELECT COUNT(id) FROM stores").fetchone()[0]
    total_patrimonio = db.execute("SELECT COUNT(id) FROM patrimonio_items").fetchone()[0]
    total_cobrancas = db.execute("SELECT COUNT(id) FROM cobranca_fichas_acerto").fetchone()[0]
    total_pagamentos = db.execute("SELECT COUNT(id) FROM contas_a_pagar_pagamentos").fetchone()[0]

    
    
    users_by_role_raw = db.execute("SELECT role, COUNT(id) as count FROM users GROUP BY role").fetchall()
    users_by_role = {row['role'].replace('_', ' ').title(): row['count'] for row in users_by_role_raw}

    
    most_active_users = db.execute("""
        SELECT username, COUNT(id) as action_count 
        FROM audit_log 
        WHERE username IS NOT NULL
        GROUP BY username 
        ORDER BY action_count DESC 
        LIMIT 5
    """).fetchall()

    
    recent_logs = db.execute("SELECT * FROM audit_log ORDER BY timestamp DESC LIMIT 5").fetchall()

    return render_template(
        'super_admin/super_admin_dashboard.html',
        total_users=total_users,
        total_stores=total_stores,
        total_patrimonio=total_patrimonio,
        total_cobrancas=total_cobrancas,
        total_pagamentos=total_pagamentos,
        users_by_role=users_by_role,
        most_active_users=most_active_users,
        recent_logs=recent_logs
    )

@app.route('/super_admin/audit_log')
@login_required
@role_required(['super_admin'])
def audit_log():
    db = get_db()
    page = request.args.get('page', 1, type=int)
    user_search = request.args.get('user_search', '').strip()
    action_filter = request.args.get('action_filter', '')
    
    base_query = "FROM audit_log WHERE 1=1"
    params = []
    
    if user_search:
        base_query += " AND username LIKE ?"
        params.append(f"%{user_search}%")
    if action_filter:
        base_query += " AND action = ?"
        params.append(action_filter)
        
    total_items = db.execute(f"SELECT COUNT(id) {base_query}", params).fetchone()[0]
    total_pages = (total_items + PER_PAGE - 1) // PER_PAGE if total_items > 0 else 1
    offset = (page - 1) * PER_PAGE

    logs = db.execute(f"SELECT * {base_query} ORDER BY timestamp DESC LIMIT ? OFFSET ?", params + [PER_PAGE, offset]).fetchall()
    distinct_actions = db.execute("SELECT DISTINCT action FROM audit_log ORDER BY action").fetchall()

    return render_template('super_admin/audit_log.html', 
                           logs=logs, 
                           distinct_actions=distinct_actions,
                           current_filters={'user': user_search, 'action': action_filter},
                           page=page,
                           total_pages=total_pages)

@app.route('/super_admin/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
@role_required(['super_admin'])
def user_edit(user_id):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        flash('Usuário não encontrado.', 'danger')
        return redirect(url_for('user_management'))

    if request.method == 'POST':
        dados_antigos = dict(user)
        dados_antigos.pop('password', None)

        role = request.form['role']
        store_id = request.form.get('store_id') or None
        can_add_users = 1 if 'can_add_users' in request.form else 0
        new_password = request.form['new_password'].strip()

        db.execute("UPDATE users SET role = ?, store_id = ?, can_add_users = ? WHERE id = ?",
                   (role, store_id, can_add_users, user_id))
        
        dados_novos = {'role': role, 'store_id': store_id, 'can_add_users': can_add_users}

        if new_password:
            if len(new_password) < 6:
                flash('A nova senha deve ter pelo menos 6 caracteres.', 'warning')
                return redirect(url_for('user_edit', user_id=user_id))
            hashed_password = generate_password_hash(new_password)
            db.execute("UPDATE users SET password = ? WHERE id = ?", (hashed_password, user_id))
            dados_novos['password'] = '******'
        
        db.commit()
        log_action('editou_usuario', target_id=user_id, target_name=user['username'], dados_antigos=dados_antigos, dados_novos=dados_novos)
        flash(f'Usuário {user["username"]} atualizado com sucesso!', 'success')
        return redirect(url_for('user_management'))

    stores = db.execute("SELECT id, name FROM stores ORDER BY name").fetchall()
    return render_template('super_admin/user_edit.html', user=user, stores=stores)

@app.route('/super_admin/users', methods=['GET', 'POST'])
@login_required
@role_required(['super_admin'])
def user_management():
    db = get_db()
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '').strip()

    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password'].strip()
        role = request.form['role']
        store_id = request.form.get('store_id') or None
        can_add_users = 1 if 'can_add_users' in request.form else 0

        if not all([username, password, role]):
            if is_ajax_request(): return jsonify({'status': 'error', 'message': 'Usuário, senha e função são obrigatórios!'}), 400
            flash('Usuário, senha e função são obrigatórios!', 'danger')
        else:
            try:
                hashed_password = generate_password_hash(password)
                cursor = db.cursor()
                cursor.execute("INSERT INTO users (username, password, role, can_add_users, store_id) VALUES (?, ?, ?, ?, ?)",
                               (username, hashed_password, role, can_add_users, store_id))
                new_user_id = cursor.lastrowid
                db.commit()
                log_action('criou_usuario', target_id=new_user_id, target_name=username)

                if is_ajax_request():
                    new_user = db.execute("SELECT u.*, s.name as store_name FROM users u LEFT JOIN stores s ON u.store_id = s.id WHERE u.id = ?", (new_user_id,)).fetchone()
                    return jsonify({'status': 'success', 'message': f'Usuário {username} criado!', 'item': dict(new_user), 'page_type': 'user_management'})
                flash(f'Usuário {username} criado com sucesso!', 'success')
            except sqlite3.IntegrityError:
                if is_ajax_request(): return jsonify({'status': 'error', 'message': f'Erro: Usuário "{username}" já existe.'}), 409
                flash(f'Erro: Usuário "{username}" já existe.', 'danger')
        return redirect(url_for('user_management'))

    params = []
    where_clauses = []
    if search_query:
        search_term = f"%{search_query}%"
        
        where_clauses.append("(u.username LIKE ? OR u.role LIKE ? OR s.name LIKE ?)")
        params.extend([search_term, search_term, search_term])
    
    base_query = "FROM users u LEFT JOIN stores s ON u.store_id = s.id"
    if where_clauses:
        base_query += " WHERE " + " AND ".join(where_clauses)

    total_items = db.execute(f"SELECT COUNT(u.id) {base_query}", params).fetchone()[0]
    total_pages = (total_items + PER_PAGE - 1) // PER_PAGE if total_items > 0 else 1
    offset = (page - 1) * PER_PAGE

    users = db.execute(f"SELECT u.*, s.name as store_name {base_query} ORDER BY u.username LIMIT ? OFFSET ?", params + [PER_PAGE, offset]).fetchall()
    stores = db.execute("SELECT id, name FROM stores ORDER BY name").fetchall()
    
    if is_ajax_request() and request.method == 'GET':
        return jsonify({
            'items': [dict(r) for r in users],
            'pagination_html': render_template('_pagination.html', page=page, total_pages=total_pages, search=search_query, request=request),
            'page_type': 'user_management',
            'script_root': request.script_root or ''
        })
        
    return render_template('super_admin/user_management.html', users=users, stores=stores, page=page, total_pages=total_pages, search=search_query)


@app.route('/super_admin/users/delete/<int:user_id>', methods=['POST'])
@login_required
@role_required(['super_admin'])
def delete_user(user_id):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if user and user['role'] == 'super_admin':
        if is_ajax_request(): return jsonify({'status': 'error', 'message': 'Não é possível excluir um Super Admin.'}), 403
        flash('Não é possível excluir um Super Admin.', 'danger')
    elif user:
        log_action('deletou_usuario', target_id=user_id, target_name=user['username'], dados_antigos=dict(user))
        db.execute("DELETE FROM users WHERE id = ?", (user_id,))
        db.commit()
        if is_ajax_request(): return jsonify({'status': 'success', 'message': 'Usuário excluído!', 'itemId': user_id})
        flash('Usuário excluído com sucesso!', 'success')
    else:
        if is_ajax_request(): return jsonify({'status': 'error', 'message': 'Usuário não encontrado.'}), 404
        flash('Usuário não encontrado.', 'danger')
    return redirect(url_for('user_management'))


@app.route('/super_admin/stores', methods=['GET', 'POST'], endpoint='manage_stores')
@login_required
@role_required(['super_admin'])
def manage_stores():
    db = get_db()
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '').strip()

    if request.method == 'POST':
        store_name = request.form.get('name', '').strip()
        selected_departments = request.form.getlist('departments')
        if not store_name or not selected_departments:
            if is_ajax_request(): return jsonify({'status': 'error', 'message': 'Nome e pelo menos um departamento são obrigatórios.'}), 400
            flash('O nome da EMPRESA e pelo menos um departamento são obrigatórios.', 'danger')
        else:
            try:
                cursor = db.cursor()
                cursor.execute("INSERT INTO stores (name) VALUES (?)", (store_name,))
                store_id = cursor.lastrowid
                for dept_role in selected_departments:
                    cursor.execute("INSERT INTO store_departments (store_id, department_role) VALUES (?, ?)", (store_id, dept_role))
                db.commit()
                log_action('criou_empresa', target_id=store_id, target_name=store_name, dados_novos={'name': store_name, 'departments': selected_departments})
                
                if is_ajax_request():
                    new_store_q = "SELECT s.id, s.name, GROUP_CONCAT(sd.department_role, ', ') as departments FROM stores s LEFT JOIN store_departments sd ON s.id = sd.store_id WHERE s.id = ? GROUP BY s.id"
                    new_store = db.execute(new_store_q, (store_id,)).fetchone()
                    return jsonify({'status': 'success', 'message': f'Empresa "{store_name}" criada!', 'item': dict(new_store), 'page_type': 'manage_stores'})
                flash(f'EMPRESA "{store_name}" criada com sucesso!', 'success')

            except sqlite3.IntegrityError:
                if is_ajax_request(): return jsonify({'status': 'error', 'message': f'Erro: A empresa "{store_name}" já existe.'}), 409
                flash(f'Erro: A EMPRESA "{store_name}" já existe.', 'danger')
        return redirect(url_for('manage_stores'))

    params = []
    where_clauses = []
    if search_query:
        search_term = f"%{search_query}%"
        search_clauses = ["s.name LIKE ?", "sd.department_role LIKE ?"]
        where_clauses.append(f"({' OR '.join(search_clauses)})")
        params.extend([search_term] * len(search_clauses))

    base_query = "FROM stores s LEFT JOIN store_departments sd ON s.id = sd.store_id"
    if where_clauses:
        base_query += " WHERE " + " AND ".join(where_clauses)
    
    count_query = f"SELECT COUNT(DISTINCT s.id) FROM stores s LEFT JOIN store_departments sd ON s.id = sd.store_id"
    if where_clauses:
        count_query += " WHERE " + " AND ".join(where_clauses)

    total_items = db.execute(count_query, params).fetchone()[0]
    total_pages = (total_items + PER_PAGE - 1) // PER_PAGE if total_items > 0 else 1
    offset = (page - 1) * PER_PAGE

    query = f"SELECT s.id, s.name, GROUP_CONCAT(sd.department_role, ', ') as departments {base_query} GROUP BY s.id, s.name ORDER BY s.name LIMIT ? OFFSET ?"
    stores_data = db.execute(query, params + [PER_PAGE, offset]).fetchall()
    
    if is_ajax_request() and request.method == 'GET':
        return jsonify({
            'items': [dict(r) for r in stores_data],
            'pagination_html': render_template('_pagination.html', page=page, total_pages=total_pages, search=search_query, request=request),
            'page_type': 'manage_stores',
            'script_root': request.script_root or ''
        })

    available_departments = ['admin_rh', 'admin_patrimonio', 'admin_contas_a_pagar', 'admin_cobranca']
    return render_template('super_admin/manage_stores.html', stores=stores_data, available_departments=available_departments, page=page, total_pages=total_pages, search=search_query)


@app.route('/super_admin/stores/delete/<int:store_id>', methods=['POST'])
@login_required
@role_required(['super_admin'])
def delete_store(store_id):
    db = get_db()
    store = db.execute("SELECT * FROM stores WHERE id = ?", (store_id,)).fetchone()
    if store:
        db.execute("UPDATE users SET store_id = NULL WHERE store_id = ?", (store_id,))
        db.execute("DELETE FROM stores WHERE id = ?", (store_id,))
        db.commit()
        log_action('deletou_empresa', target_id=store_id, target_name=store['name'], dados_antigos=dict(store))
        if is_ajax_request():
            return jsonify({'status': 'success', 'message': 'Empresa excluída!', 'itemId': store_id})
        flash('EMPRESA excluída com sucesso!', 'success')
    else:
        if is_ajax_request():
            return jsonify({'status': 'error', 'message': 'Empresa não encontrada.'}), 404
        flash('EMPRESA não encontrada.', 'danger')
    return redirect(url_for('manage_stores'))


@app.route('/super_admin/manage_marcas', methods=['GET', 'POST'])
@login_required
@role_required(['super_admin'])
def manage_marcas():
    db = get_db()
    if request.method == 'POST':
        nome_marca = request.form.get('nome_marca')
        tipo_id = request.form.get('tipo_id')
        if nome_marca and tipo_id:
            db.execute("INSERT INTO marcas (nome, tipo_id) VALUES (?, ?)", (nome_marca, tipo_id))
            db.commit()
            flash('Marca adicionada com sucesso!', 'success')
        else:
            flash('Nome da marca e tipo são obrigatórios.', 'danger')
        return redirect(url_for('manage_marcas'))
    
    tipos = db.execute("SELECT * FROM tipos_equipamento").fetchall()
    marcas = db.execute("""
        SELECT m.id, m.nome, t.nome as tipo_nome 
        FROM marcas m 
        JOIN tipos_equipamento t ON m.tipo_id = t.id 
        ORDER BY t.nome, m.nome
    """).fetchall()
    
    return render_template('super_admin/manage_marcas.html', tipos=tipos, marcas=marcas)

@app.route('/super_admin/manage_attributes')
@login_required
@role_required(['super_admin'])
def manage_attributes():
    db = get_db()
    tipos = db.execute("SELECT * FROM tipos_equipamento ORDER BY nome").fetchall()
    marcas = db.execute("""
        SELECT m.id, m.nome, t.nome as tipo_nome
        FROM marcas m 
        JOIN tipos_equipamento t ON m.tipo_id = t.id 
        ORDER BY t.nome, m.nome
    """).fetchall()
    tamanhos = db.execute("SELECT * FROM tamanhos ORDER BY nome").fetchall()
    
    return render_template('super_admin/manage_attributes.html', tipos=tipos, marcas=marcas, tamanhos=tamanhos)


@app.route('/super_admin/tipos/add', methods=['POST'])
@login_required
@role_required(['super_admin'])
def add_tipo():
    nome_tipo = request.form.get('nome_tipo')
    if nome_tipo:
        try:
            db = get_db()
            db.execute("INSERT INTO tipos_equipamento (nome) VALUES (?)", (nome_tipo,))
            db.commit()
            flash(f'Tipo "{nome_tipo}" adicionado com sucesso!', 'success')
        except sqlite3.IntegrityError:
            flash(f'O tipo "{nome_tipo}" já existe.', 'danger')
    else:
        flash('O nome do tipo não pode ser vazio.', 'danger')
    return redirect(url_for('manage_attributes'))

@app.route('/super_admin/marcas/add', methods=['POST'])
@login_required
@role_required(['super_admin'])
def add_marca():
    nome_marca = request.form.get('nome_marca')
    tipo_id = request.form.get('tipo_id')
    if nome_marca and tipo_id:
        db = get_db()
        db.execute("INSERT INTO marcas (nome, tipo_id) VALUES (?, ?)", (nome_marca, tipo_id))
        db.commit()
        flash('Marca adicionada com sucesso!', 'success')
    else:
        flash('Nome da marca e tipo são obrigatórios.', 'danger')
    return redirect(url_for('manage_attributes'))

@app.route('/super_admin/tamanhos/add', methods=['POST'])
@login_required
@role_required(['super_admin'])
def add_tamanho():
    nome_tamanho = request.form.get('nome_tamanho')
    if nome_tamanho:
        try:
            db = get_db()
            db.execute("INSERT INTO tamanhos (nome) VALUES (?)", (nome_tamanho,))
            db.commit()
            flash(f'Tamanho "{nome_tamanho}" adicionado com sucesso!', 'success')
        except sqlite3.IntegrityError:
            flash(f'O tamanho "{nome_tamanho}" já existe.', 'danger')
    else:
        flash('O nome do tamanho não pode ser vazio.', 'danger')
    return redirect(url_for('manage_attributes'))


@app.route('/super_admin/tipos/delete/<int:tipo_id>', methods=['POST'])
@login_required
@role_required(['super_admin'])
def delete_tipo(tipo_id):
    db = get_db()
    
    db.execute("DELETE FROM tipos_equipamento WHERE id = ?", (tipo_id,))
    db.commit()
    flash('Tipo removido com sucesso.', 'success')
    return redirect(url_for('manage_attributes'))

@app.route('/super_admin/marcas/delete/<int:marca_id>', methods=['POST'])
@login_required
@role_required(['super_admin'])
def delete_marca(marca_id):
    db = get_db()
    db.execute("DELETE FROM marcas WHERE id = ?", (marca_id,))
    db.commit()
    flash('Marca removida com sucesso!', 'success')
    return redirect(url_for('manage_attributes'))

@app.route('/super_admin/tamanhos/delete/<int:tamanho_id>', methods=['POST'])
@login_required
@role_required(['super_admin'])
def delete_tamanho(tamanho_id):
    db = get_db()
    db.execute("DELETE FROM tamanhos WHERE id = ?", (tamanho_id,))
    db.commit()
    flash('Tamanho removido com sucesso!', 'success')
    return redirect(url_for('manage_attributes'))


@app.route('/patrimonio/dashboard', methods=['GET', 'POST'])
@login_required
@role_required(['super_admin', 'admin_patrimonio'])
def patrimonio_dashboard():
    db = get_db()
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '').strip()

    if request.method == 'POST':
        codigo_cliente = request.form['codigo_cliente']
        nome_cliente = request.form['nome_cliente']
        numero_caixa = request.form['numero_caixa']
        codigo_patrimonio = request.form['codigo_patrimonio']
        tipo_id = request.form.get('tipo_id')
        marca_id = request.form.get('marca_id')
        tamanho = request.form.get('tamanho')
        store_id = session.get('store_id') if session.get('role') != 'super_admin' else request.form.get('store_id')

        if not all([codigo_cliente, nome_cliente, numero_caixa, codigo_patrimonio, tipo_id, marca_id, store_id]):
             flash('Todos os campos obrigatórios, incluindo o Código do Patrimônio, devem ser preenchidos.', 'danger')
             return redirect(url_for('patrimonio_dashboard'))

        with db:
            cliente = db.execute("SELECT * FROM clientes WHERE codigo_cliente = ? AND store_id = ?", (codigo_cliente, store_id)).fetchone()
            if cliente:
                cliente_id = cliente['id']
                if cliente['numero_caixa'] != numero_caixa:
                    db.execute("UPDATE clientes SET numero_caixa = ? WHERE id = ?", (numero_caixa, cliente_id))
            else:
                cursor = db.execute("INSERT INTO clientes (codigo_cliente, nome_cliente, numero_caixa, store_id) VALUES (?, ?, ?, ?)",
                                    (codigo_cliente, nome_cliente, numero_caixa, store_id))
                cliente_id = cursor.lastrowid
            
            try:
                db.execute("INSERT INTO patrimonio_items (cliente_id, codigo_patrimonio, tipo_id, marca_id, tamanho) VALUES (?, ?, ?, ?, ?)",
                           (cliente_id, codigo_patrimonio, tipo_id, marca_id, tamanho))
                flash('Patrimônio adicionado com sucesso!', 'success')
            except sqlite3.IntegrityError:
                flash('Erro: O código de patrimônio já existe para este cliente.', 'danger')
        
        return redirect(url_for('patrimonio_dashboard'))

    params = []
    where_clauses = []
    if session['role'] != 'super_admin':
        where_clauses.append("c.store_id = ?")
        params.append(session.get('store_id'))

    if search_query:
        search_term = f"%{search_query}%"
        where_clauses.append("(c.codigo_cliente LIKE ? OR c.nome_cliente LIKE ? OR c.numero_caixa LIKE ?)")
        params.extend([search_term, search_term, search_term])

    base_query = "FROM clientes c LEFT JOIN stores s ON c.store_id = s.id"
    if where_clauses:
        base_query += " WHERE " + " AND ".join(where_clauses)

    total_items = db.execute(f"SELECT COUNT(c.id) {base_query}", params).fetchone()[0]
    total_pages = (total_items + PER_PAGE - 1) // PER_PAGE if total_items > 0 else 1
    offset = (page - 1) * PER_PAGE
    
    clientes_raw = db.execute(f"SELECT c.*, s.name as store_name {base_query} ORDER BY c.nome_cliente ASC LIMIT ? OFFSET ?", params + [PER_PAGE, offset]).fetchall()
    
    clientes_list = []
    for cliente in clientes_raw:
        cliente_dict = dict(cliente)
        count = db.execute("SELECT COUNT(id) FROM patrimonio_items WHERE cliente_id = ?", (cliente_dict['id'],)).fetchone()[0]
        cliente_dict['patrimonio_count'] = count
        clientes_list.append(cliente_dict)

    tipos = db.execute("SELECT * FROM tipos_equipamento ORDER BY nome").fetchall()
    stores = db.execute("SELECT * FROM stores ORDER BY name").fetchall() if session['role'] == 'super_admin' else []
    tamanhos = db.execute("SELECT * FROM tamanhos ORDER BY nome").fetchall()
    stores = db.execute("SELECT * FROM stores ORDER BY name").fetchall() if session['role'] == 'super_admin' else []

    return render_template('patrimonio/patrimonio_dashboard.html', clientes=clientes_list, tipos=tipos, tamanhos=tamanhos, stores=stores, page=page, total_pages=total_pages, search=search_query)

@app.route('/patrimonio/export_excel')
@login_required
@role_required(['super_admin', 'admin_patrimonio'])
def patrimonio_export_excel():
    db = get_db()
    
    
    query = """
        SELECT 
            s.name as nome_empresa, c.codigo_cliente, c.nome_cliente, c.numero_caixa,
            pi.codigo_patrimonio, t.nome as tipo, m.nome as marca, pi.tamanho
        FROM clientes c
        JOIN patrimonio_items pi ON c.id = pi.cliente_id
        JOIN tipos_equipamento t ON pi.tipo_id = t.id
        JOIN marcas m ON pi.marca_id = m.id
        LEFT JOIN stores s ON c.store_id = s.id
    """
    params = []

    if session['role'] != 'super_admin':
        query += " WHERE c.store_id = ?"
        params.append(session.get('store_id'))

    query += " ORDER BY s.name, c.nome_cliente, pi.codigo_patrimonio"
    
    items = db.execute(query, params).fetchall()
    
    
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Relatório de Patrimônios'

    
    cabecalho = ['Empresa', 'Código Cliente', 'Nome Cliente', 'Nº Caixa', 'Código Patrimônio', 'Tipo', 'Marca', 'Tamanho']
    sheet.append(cabecalho)
    
    
    for item in items:
        sheet.append([
            item['nome_empresa'] or 'N/A',
            item['codigo_cliente'],
            item['nome_cliente'],
            item['numero_caixa'],
            item['codigo_patrimonio'],
            item['tipo'],
            item['marca'],
            item['tamanho'] or 'N/A'
        ])
        
    
    
    
    full_range = f"A1:{openpyxl.utils.get_column_letter(sheet.max_column)}{sheet.max_row}"
    
    
    tab = Table(displayName="TabelaPatrimonios", ref=full_range)

    
    
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    
    tab.tableStyleInfo = style
    
    
    sheet.add_table(tab)

    
    sheet.freeze_panes = 'A2'

    
    for col in sheet.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    
    workbook.save(output)
    output.seek(0)
    
    return Response(output,
                   mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                   headers={"Content-Disposition": "attachment;filename=Relatorio_Patrimonios.xlsx"})

@app.route('/patrimonio/delete/<int:cliente_id>', methods=['POST'])
@login_required
@role_required(['super_admin', 'admin_patrimonio'])
def patrimonio_delete(cliente_id):
    db = get_db()
    cliente = db.execute("SELECT * FROM clientes WHERE id = ?", (cliente_id,)).fetchone()
    if not cliente or (session['role'] != 'super_admin' and cliente.get('store_id') != session.get('store_id')):
        flash('Cliente não encontrado ou sem permissão para excluir.', 'danger')
    else:
        with db:
            log_action('deletou_cliente', target_id=cliente_id, target_name=cliente['codigo_cliente'], dados_antigos=dict(cliente))
            db.execute("DELETE FROM patrimonio_items WHERE cliente_id = ?", (cliente_id,))
            db.execute("DELETE FROM clientes WHERE id = ?", (cliente_id,))
        flash('Cliente e todos os seus patrimônios foram excluídos com sucesso!', 'success')
    return redirect(url_for('patrimonio_dashboard'))

@app.route('/patrimonio/edit/<int:cliente_id>', methods=['GET', 'POST'])
@login_required
@role_required(['super_admin', 'admin_patrimonio'])
def patrimonio_edit(cliente_id):
    db = get_db()
    cliente = db.execute("SELECT * FROM clientes WHERE id = ?", (cliente_id,)).fetchone()
    if not cliente or (session['role'] != 'super_admin' and cliente.get('store_id') != session.get('store_id')):
        flash('Cliente não encontrado ou sem permissão.', 'danger')
        return redirect(url_for('patrimonio_dashboard'))

    if request.method == 'POST':
        db.execute("UPDATE clientes SET codigo_cliente = ?, nome_cliente = ?, numero_caixa = ? WHERE id = ?",
                   (request.form['codigo_cliente'], request.form['nome_cliente'], request.form['numero_caixa'], cliente_id))
        db.commit()
        flash('Dados do cliente atualizados com sucesso!', 'success')
        
        return redirect(url_for('patrimonio_edit', cliente_id=cliente_id))

    patrimonios = db.execute("""
        SELECT pi.id, pi.codigo_patrimonio, t.nome as tipo, m.nome as marca, pi.tamanho
        FROM patrimonio_items pi
        JOIN tipos_equipamento t ON pi.tipo_id = t.id
        JOIN marcas m ON pi.marca_id = m.id
        WHERE pi.cliente_id = ? ORDER BY pi.codigo_patrimonio
    """, (cliente_id,)).fetchall()
    
    return render_template('patrimonio/patrimonio_edit.html', cliente=cliente, patrimonios=patrimonios)

@app.route('/patrimonio/item/delete/<int:item_id>', methods=['POST'])
@login_required
@role_required(['super_admin', 'admin_patrimonio'])
def delete_patrimonio_item(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM patrimonio_items WHERE id = ?", (item_id,)).fetchone()
    if item:
        cliente_id = item['cliente_id']
        cliente = db.execute("SELECT store_id FROM clientes WHERE id = ?", (cliente_id,)).fetchone()
        if session['role'] == 'super_admin' or cliente['store_id'] == session.get('store_id'):
            db.execute("DELETE FROM patrimonio_items WHERE id = ?", (item_id,))
            db.commit()
            flash('Patrimônio removido com sucesso!', 'success')
            return redirect(url_for('patrimonio_edit', cliente_id=cliente_id))
    flash('Item não encontrado ou sem permissão.', 'danger')
    return redirect(url_for('patrimonio_dashboard'))

@app.route('/patrimonio/relatorios', methods=['GET'])
@login_required
@role_required(['super_admin', 'admin_patrimonio'])
def patrimonio_relatorios():
    db = get_db()
    
    tipo_id = request.args.get('tipo_id', type=int)
    marca_id = request.args.get('marca_id', type=int)
    tamanho = request.args.get('tamanho', '')
    store_id = request.args.get('store_id', type=int)

    query = """
        SELECT t.nome as tipo, m.nome as marca, pi.tamanho, s.name as store_name, COUNT(pi.id) as total
        FROM patrimonio_items pi
        JOIN clientes c ON pi.cliente_id = c.id
        JOIN tipos_equipamento t ON pi.tipo_id = t.id
        JOIN marcas m ON pi.marca_id = m.id
        LEFT JOIN stores s ON c.store_id = s.id
        WHERE 1=1
    """
    params = []

    if session['role'] != 'super_admin':
        query += " AND c.store_id = ?"
        params.append(session.get('store_id'))
    elif store_id:
        query += " AND c.store_id = ?"
        params.append(store_id)

    if tipo_id:
        query += " AND pi.tipo_id = ?"
        params.append(tipo_id)
    if marca_id:
        query += " AND pi.marca_id = ?"
        params.append(marca_id)
    if tamanho:
        query += " AND pi.tamanho = ?"
        params.append(tamanho)

    query += " GROUP BY s.name, t.nome, m.nome, pi.tamanho ORDER BY s.name, t.nome, m.nome"
    
    resultados = db.execute(query, params).fetchall()

    tipos = db.execute("SELECT * FROM tipos_equipamento ORDER BY nome").fetchall()
    marcas = db.execute("SELECT * FROM marcas ORDER BY nome").fetchall()
    stores = db.execute("SELECT * FROM stores ORDER BY name").fetchall()
    
    return render_template('patrimonio/relatorios.html', 
                           resultados=resultados, 
                           tipos=tipos, 
                           marcas=marcas,
                           stores=stores,
                           current_filters={'tipo_id': tipo_id, 'marca_id': marca_id, 'tamanho': tamanho, 'store_id': store_id})


@app.route('/rh/dashboard')
@login_required
@role_required(['super_admin', 'admin_rh'])
def rh_dashboard():
    return render_template('rh/rh_dashboard.html')



@app.route('/contas_a_pagar/dashboard', methods=['GET', 'POST'])
@login_required
@role_required(['super_admin', 'admin_contas_a_pagar'])
def contas_a_pagar_dashboard():
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '').strip()
    db = get_db()

    if request.method == 'POST':
        store_id = session.get('store_id') if session['role'] != 'super_admin' else request.form.get('store_id')
        if session['role'] == 'super_admin' and not store_id:
            if is_ajax_request(): return jsonify({'status': 'error', 'message': 'Como Super Admin, você deve selecionar uma EMPRESA.'}), 400
            flash('Como Super Admin, você deve selecionar uma EMPRESA.', 'danger')
            return redirect(url_for('contas_a_pagar_dashboard'))
        
        dados_novos = {'pagamento_data_inicio': request.form['pagamento_data_inicio'], 'pagamento_data_fim': request.form['pagamento_data_fim'], 'caixa': request.form['caixa'], 'store_id': store_id}
        cursor = db.cursor()
        cursor.execute("INSERT INTO contas_a_pagar_pagamentos (pagamento_data_inicio, pagamento_data_fim, caixa, store_id) VALUES (?, ?, ?, ?)", list(dados_novos.values()))
        new_id = cursor.lastrowid
        db.commit()
        log_action('adicionou_pagamento', target_id=new_id, target_name=f"Caixa {dados_novos['caixa']}", dados_novos=dados_novos)

        if is_ajax_request():
            new_item = db.execute("SELECT p.*, s.name as store_name FROM contas_a_pagar_pagamentos p LEFT JOIN stores s ON p.store_id = s.id WHERE p.id = ?", (new_id,)).fetchone()
            item_dict = dict(new_item)
            item_dict['pagamento_data_inicio'] = datetime.strptime(item_dict['pagamento_data_inicio'], '%Y-%m-%d').strftime('%d/%m/%Y')
            item_dict['pagamento_data_fim'] = datetime.strptime(item_dict['pagamento_data_fim'], '%Y-%m-%d').strftime('%d/%m/%Y')
            return jsonify({'status': 'success', 'message': 'Pagamento adicionado!', 'item': item_dict, 'page_type': 'contas_a_pagar_pagamentos'})
        flash('Pagamento adicionado com sucesso!', 'success')
        return redirect(url_for('contas_a_pagar_dashboard'))

    params = []
    where_clauses = []

    if session['role'] != 'super_admin':
        where_clauses.append("p.store_id = ?")
        params.append(session.get('store_id'))

    if search_query:
        search_term = f"%{search_query}%"
        search_clauses_list = ["p.caixa LIKE ?", "p.pagamento_data_inicio LIKE ?", "p.pagamento_data_fim LIKE ?"]
        if session['role'] == 'super_admin':
            search_clauses_list.append("s.name LIKE ?")
        
        where_clauses.append(f"({' OR '.join(search_clauses_list)})")
        params.extend([search_term] * len(search_clauses_list))

    base_query = "FROM contas_a_pagar_pagamentos p LEFT JOIN stores s ON p.store_id = s.id"
    if where_clauses:
        base_query += " WHERE " + " AND ".join(where_clauses)


    total_items = db.execute(f"SELECT COUNT(p.id) {base_query}", params).fetchone()[0]
    total_pages = (total_items + PER_PAGE - 1) // PER_PAGE if total_items > 0 else 1
    offset = (page - 1) * PER_PAGE
    items = db.execute(f"SELECT p.*, s.name as store_name {base_query} ORDER BY p.id DESC LIMIT ? OFFSET ?", params + [PER_PAGE, offset]).fetchall()
    
    items_formatted = []
    for item in items:
        item_dict = dict(item)
        try:
            item_dict['pagamento_data_inicio'] = datetime.strptime(item_dict['pagamento_data_inicio'], '%Y-%m-%d').strftime('%d/%m/%Y')
            item_dict['pagamento_data_fim'] = datetime.strptime(item_dict['pagamento_data_fim'], '%Y-%m-%d').strftime('%d/%m/%Y')
        except (ValueError, TypeError): pass
        items_formatted.append(item_dict)

    if is_ajax_request() and request.method == 'GET':
        return jsonify({
            'items': items_formatted,
            'pagination_html': render_template('_pagination.html', page=page, total_pages=total_pages, search=search_query, request=request),
            'page_type': 'contas_a_pagar_pagamentos',
            'script_root': request.script_root or ''
        })

    stores = db.execute("SELECT id, name FROM stores ORDER BY name").fetchall() if session['role'] == 'super_admin' else []
    return render_template('contas_a_pagar/contas_a_pagar_dashboard.html', pagamentos=items_formatted, stores=stores, page=page, total_pages=total_pages, search=search_query)


@app.route('/contas_a_pagar/pagamentos/delete/<int:item_id>', methods=['POST'])
@login_required
@role_required(['super_admin', 'admin_contas_a_pagar'])
def contas_a_pagar_pagamentos_delete(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM contas_a_pagar_pagamentos WHERE id = ?", (item_id,)).fetchone()
    if not item or (session['role'] != 'super_admin' and item['store_id'] != session.get('store_id')):
        if is_ajax_request(): return jsonify({'status': 'error', 'message': 'Registro não encontrado ou sem permissão.'}), 403
        flash('Registro não encontrado ou sem permissão.', 'danger')
    else:
        log_action('deletou_pagamento', target_id=item_id, target_name=f"Caixa {item['caixa']}", dados_antigos=dict(item))
        db.execute("DELETE FROM contas_a_pagar_pagamentos WHERE id = ?", (item_id,))
        db.commit()
        if is_ajax_request(): return jsonify({'status': 'success', 'message': 'Registro excluído!', 'itemId': item_id})
        flash('Registro excluído com sucesso!', 'success')
    return redirect(url_for('contas_a_pagar_dashboard'))


@app.route('/contas_a_pagar/pagamentos/edit/<int:item_id>', methods=['GET', 'POST'])
@login_required
@role_required(['super_admin', 'admin_contas_a_pagar'])
def contas_a_pagar_pagamentos_edit(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM contas_a_pagar_pagamentos WHERE id = ?", (item_id,)).fetchone()
    if not item or (session['role'] != 'super_admin' and item['store_id'] != session.get('store_id')):
        flash('Registro não encontrado ou sem permissão.', 'danger')
        return redirect(url_for('contas_a_pagar_dashboard'))

    if request.method == 'POST':
        dados_antigos = dict(item)
        dados_novos = {'pagamento_data_inicio': request.form['pagamento_data_inicio'], 'pagamento_data_fim': request.form['pagamento_data_fim'], 'caixa': request.form['caixa']}
        db.execute("UPDATE contas_a_pagar_pagamentos SET pagamento_data_inicio = ?, pagamento_data_fim = ?, caixa = ? WHERE id = ?", 
                   (dados_novos['pagamento_data_inicio'], dados_novos['pagamento_data_fim'], dados_novos['caixa'], item_id))
        db.commit()
        log_action('editou_pagamento', target_id=item_id, target_name=f"Caixa {dados_novos['caixa']}", dados_antigos=dados_antigos, dados_novos=dados_novos)
        flash('Registro atualizado com sucesso!', 'success')
        return redirect(url_for('contas_a_pagar_dashboard'))
    return render_template('contas_a_pagar/contas_a_pagar_pagamentos_edit.html', item=item)



@app.route('/cobranca/dashboard', methods=['GET', 'POST'])
@login_required
@role_required(['super_admin', 'admin_cobranca'])
def cobranca_dashboard():
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '').strip()
    db = get_db()

    if request.method == 'POST':
        store_id = session.get('store_id') if session['role'] != 'super_admin' else request.form.get('store_id')
        if session['role'] == 'super_admin' and not store_id:
            if is_ajax_request(): return jsonify({'status': 'error', 'message': 'Como Super Admin, você deve selecionar uma EMPRESA.'}), 400
            flash('Como Super Admin, você deve selecionar uma EMPRESA.', 'danger')
            return redirect(url_for('cobranca_dashboard'))
            
        try:
            dados_novos = {'ficha_acerto': request.form['ficha_acerto'], 'caixa': request.form['caixa'], 'range_cliente_inicio': int(request.form['range_cliente_inicio']), 'range_cliente_fim': int(request.form['range_cliente_fim']), 'store_id': store_id, 'order_position': 0}
            cursor = db.cursor()
            cursor.execute("INSERT INTO cobranca_fichas_acerto (ficha_acerto, caixa, range_cliente_inicio, range_cliente_fim, store_id, order_position) VALUES (?, ?, ?, ?, ?, ?)", list(dados_novos.values()))
            new_id = cursor.lastrowid
            db.commit()
            log_action('adicionou_ficha_acerto', target_id=new_id, target_name=f"Ficha {dados_novos['ficha_acerto']}", dados_novos=dados_novos)

            if is_ajax_request():
                new_item = db.execute("SELECT f.*, s.name as store_name FROM cobranca_fichas_acerto f LEFT JOIN stores s ON f.store_id = s.id WHERE f.id = ?", (new_id,)).fetchone()
                return jsonify({'status': 'success', 'message': 'Ficha de Acerto adicionada!', 'item': dict(new_item), 'page_type': 'cobranca'})
            flash('Ficha de Acerto adicionada com sucesso!', 'success')
        except ValueError:
            if is_ajax_request(): return jsonify({'status': 'error', 'message': 'Os campos de range de cliente devem ser números.'}), 400
            flash('Os campos de range de cliente devem ser números.', 'danger')
        return redirect(url_for('cobranca_dashboard'))

    params = []
    where_clauses = []

    if session['role'] != 'super_admin':
        where_clauses.append("f.store_id = ?")
        params.append(session.get('store_id'))

    if search_query:
        search_term = f"%{search_query}%"
        search_clauses_list = ["f.ficha_acerto LIKE ?", "f.caixa LIKE ?", "f.range_cliente_inicio LIKE ?", "f.range_cliente_fim LIKE ?"]
        if session['role'] == 'super_admin':
            search_clauses_list.append("s.name LIKE ?")
        
        where_clauses.append(f"({' OR '.join(search_clauses_list)})")
        params.extend([search_term] * len(search_clauses_list))

    base_query = "FROM cobranca_fichas_acerto f LEFT JOIN stores s ON f.store_id = s.id"
    if where_clauses:
        base_query += " WHERE " + " AND ".join(where_clauses)


    total_items = db.execute(f"SELECT COUNT(f.id) {base_query}", params).fetchone()[0]
    total_pages = (total_items + PER_PAGE - 1) // PER_PAGE if total_items > 0 else 1
    offset = (page - 1) * PER_PAGE
    items = db.execute(f"SELECT f.*, s.name as store_name {base_query} ORDER BY CAST(f.caixa as INTEGER) ASC, f.range_cliente_inicio ASC LIMIT ? OFFSET ?", params + [PER_PAGE, offset]).fetchall()
    
    if is_ajax_request() and request.method == 'GET':
        return jsonify({
            'items': [dict(r) for r in items],
            'pagination_html': render_template('_pagination.html', page=page, total_pages=total_pages, search=search_query, request=request),
            'page_type': 'cobranca',
            'script_root': request.script_root or ''
        })
    
    stores = db.execute("SELECT id, name FROM stores ORDER BY name").fetchall() if session['role'] == 'super_admin' else []
    return render_template('cobranca/cobranca_dashboard.html', fichas_acerto=items, stores=stores, page=page, total_pages=total_pages, search=search_query)

@app.route('/cobranca/export_excel')
@login_required
@role_required(['super_admin', 'admin_cobranca'])
def cobranca_export_excel():
    db = get_db()
    query = "SELECT f.ficha_acerto, f.caixa, f.range_cliente_inicio, f.range_cliente_fim, s.name as store_name FROM cobranca_fichas_acerto f LEFT JOIN stores s ON f.store_id = s.id"
    params = []
    if session['role'] != 'super_admin':
        query += " WHERE f.store_id = ?"
        params.append(session.get('store_id'))
    
    query += " ORDER BY CAST(f.caixa as INTEGER) ASC, f.range_cliente_inicio ASC"
    items = db.execute(query, params).fetchall()
    
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Fichas de Acerto"

    cabecalho = ['Empresa', 'Ficha de Acerto', 'Caixa', 'Range Cliente Início', 'Range Cliente Fim']
    sheet.append(cabecalho)
    
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for item in items:
        sheet.append([
            item['store_name'] or 'N/A',
            item['ficha_acerto'], 
            item['caixa'], 
            item['range_cliente_inicio'], 
            item['range_cliente_fim']
        ])
    
    
    tabela = Table(displayName="TabelaCobranca", ref=f"A1:{openpyxl.utils.get_column_letter(sheet.max_column)}{sheet.max_row}")
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tabela.tableStyleInfo = estilo
    sheet.add_table(tabela)

    for col in sheet.columns:
        max_length = 0
        column = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
        
    workbook.save(output)
    output.seek(0)
    
    return Response(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition":"attachment;filename=cobranca_fichas_acerto.xlsx"})

@app.route('/cobranca/fichas_acerto/delete/<int:item_id>', methods=['POST'])
@login_required
@role_required(['super_admin', 'admin_cobranca'])
def cobranca_fichas_acerto_delete(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM cobranca_fichas_acerto WHERE id = ?", (item_id,)).fetchone()
    if not item or (session['role'] != 'super_admin' and item['store_id'] != session.get('store_id')):
        if is_ajax_request(): return jsonify({'status': 'error', 'message': 'Ficha de Acerto não encontrada ou sem permissão.'}), 403
        flash('Ficha de Acerto não encontrada ou sem permissão.', 'danger')
    else:
        log_action('deletou_ficha_acerto', target_id=item_id, target_name=f"Ficha {item['ficha_acerto']}", dados_antigos=dict(item))
        db.execute("DELETE FROM cobranca_fichas_acerto WHERE id = ?", (item_id,))
        db.commit()
        if is_ajax_request(): return jsonify({'status': 'success', 'message': 'Ficha de Acerto excluída!', 'itemId': item_id})
        flash('Ficha de Acerto excluída com sucesso!', 'success')
    return redirect(url_for('cobranca_dashboard'))


@app.route('/cobranca/fichas_acerto/edit/<int:item_id>', methods=['GET', 'POST'])
@login_required
@role_required(['super_admin', 'admin_cobranca'])
def cobranca_fichas_acerto_edit(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM cobranca_fichas_acerto WHERE id = ?", (item_id,)).fetchone()
    if not item or (session['role'] != 'super_admin' and item['store_id'] != session.get('store_id')):
        flash('Ficha de Acerto não encontrada ou sem permissão.', 'danger')
        return redirect(url_for('cobranca_dashboard'))

    if request.method == 'POST':
        try:
            dados_antigos = dict(item)
            dados_novos = {'ficha_acerto': request.form['ficha_acerto'], 'caixa': request.form['caixa'], 'range_cliente_inicio': int(request.form['range_cliente_inicio']), 'range_cliente_fim': int(request.form['range_cliente_fim'])}
            db.execute("UPDATE cobranca_fichas_acerto SET ficha_acerto = ?, caixa = ?, range_cliente_inicio = ?, range_cliente_fim = ? WHERE id = ?", 
                       (dados_novos['ficha_acerto'], dados_novos['caixa'], dados_novos['range_cliente_inicio'], dados_novos['range_cliente_fim'], item_id))
            db.commit()
            log_action('editou_ficha_acerto', target_id=item_id, target_name=f"Ficha {dados_novos['ficha_acerto']}", dados_antigos=dados_antigos, dados_novos=dados_novos)
            flash('Ficha de Acerto atualizada com sucesso!', 'success')
        except ValueError:
            flash('Os campos de range de cliente devem ser números.', 'danger')
        return redirect(url_for('cobranca_dashboard'))
    return render_template('cobranca/cobranca_fichas_acerto_edit.html', item=item)


@app.route('/contas_a_pagar/documentos_diversos', methods=['GET', 'POST'])
@login_required
@role_required(['super_admin', 'admin_contas_a_pagar'])
def documentos_diversos_dashboard():
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '').strip()
    db = get_db()

    if request.method == 'POST':
        store_id = session.get('store_id') if session['role'] != 'super_admin' else request.form.get('store_id')
        if session['role'] == 'super_admin' and not store_id:
            if is_ajax_request(): return jsonify({'status': 'error', 'message': 'Como Super Admin, você deve selecionar uma EMPRESA.'}), 400
            flash('Como Super Admin, você deve selecionar uma EMPRESA.', 'danger')
            return redirect(url_for('documentos_diversos_dashboard'))

        dados_novos = {'numero_caixa': request.form['numero_caixa'], 'store_id': store_id}
        cursor = db.cursor()
        cursor.execute("INSERT INTO contas_a_pagar_diversos (numero_caixa, store_id) VALUES (?, ?)", list(dados_novos.values()))
        new_id = cursor.lastrowid
        db.commit()
        log_action('adicionou_documento_diverso', target_id=new_id, target_name=f"Caixa {dados_novos['numero_caixa']}", dados_novos=dados_novos)

        if is_ajax_request():
            new_item = db.execute("SELECT d.*, s.name as store_name FROM contas_a_pagar_diversos d LEFT JOIN stores s ON d.store_id = s.id WHERE d.id = ?", (new_id,)).fetchone()
            return jsonify({'status': 'success', 'message': 'Documento adicionado!', 'item': dict(new_item), 'page_type': 'contas_a_pagar_diversos'})
        flash('Documento diverso adicionado com sucesso!', 'success')
        return redirect(url_for('documentos_diversos_dashboard'))

    params = []
    where_clauses = []

    if session['role'] != 'super_admin':
        where_clauses.append("d.store_id = ?")
        params.append(session.get('store_id'))

    if search_query:
        search_term = f"%{search_query}%"
        search_clauses_list = ["d.numero_caixa LIKE ?"]
        if session['role'] == 'super_admin':
            search_clauses_list.append("s.name LIKE ?")
        
        where_clauses.append(f"({' OR '.join(search_clauses_list)})")
        params.extend([search_term] * len(search_clauses_list))

    base_query = "FROM contas_a_pagar_diversos d LEFT JOIN stores s ON d.store_id = s.id"
    if where_clauses:
        base_query += " WHERE " + " AND ".join(where_clauses)

    total_items = db.execute(f"SELECT COUNT(d.id) {base_query}", params).fetchone()[0]
    total_pages = (total_items + PER_PAGE - 1) // PER_PAGE if total_items > 0 else 1
    offset = (page - 1) * PER_PAGE
    items = db.execute(f"SELECT d.*, s.name as store_name {base_query} ORDER BY d.id DESC LIMIT ? OFFSET ?", params + [PER_PAGE, offset]).fetchall()
    
    if is_ajax_request() and request.method == 'GET':
        return jsonify({
            'items': [dict(r) for r in items],
            'pagination_html': render_template('_pagination.html', page=page, total_pages=total_pages, request=request, search=search_query),
            'page_type': 'contas_a_pagar_diversos',
            'script_root': request.script_root or ''
        })
    
    stores = db.execute("SELECT id, name FROM stores ORDER BY name").fetchall() if session['role'] == 'super_admin' else []
    return render_template('contas_a_pagar/documentos_diversos_dashboard.html', documentos_diversos=items, stores=stores, page=page, total_pages=total_pages, search=search_query)

@app.route('/contas_a_pagar/pagamentos/export_excel')
@login_required
@role_required(['super_admin', 'admin_contas_a_pagar'])
def contas_a_pagar_pagamentos_export_excel():
    db = get_db()
    query = "SELECT p.pagamento_data_inicio, p.pagamento_data_fim, p.caixa, s.name as store_name FROM contas_a_pagar_pagamentos p LEFT JOIN stores s ON p.store_id = s.id"
    params = []
    if session['role'] != 'super_admin':
        query += " WHERE p.store_id = ?"
        params.append(session.get('store_id'))
    
    items = db.execute(query + " ORDER BY p.id DESC", params).fetchall()

    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Pagamentos"
    cabecalho = ['Empresa', 'Data Início', 'Data Fim', 'Caixa']
    sheet.append(cabecalho)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for item in items:
        sheet.append([item['store_name'] or 'N/A', item['pagamento_data_inicio'], item['pagamento_data_fim'], item['caixa']])

    tabela = Table(displayName="TabelaPagamentos", ref=f"A1:{openpyxl.utils.get_column_letter(sheet.max_column)}{sheet.max_row}")
    estilo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tabela.tableStyleInfo = estilo
    sheet.add_table(tabela)

    for col in sheet.columns:
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = 20

    workbook.save(output)
    output.seek(0)
    return Response(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition":"attachment;filename=contas_a_pagar_pagamentos.xlsx"})

@app.route('/contas_a_pagar/documentos_diversos/export_excel')
@login_required
@role_required(['super_admin', 'admin_contas_a_pagar'])
def contas_a_pagar_diversos_export_excel():
    db = get_db()
    query = "SELECT d.numero_caixa, s.name as store_name FROM contas_a_pagar_diversos d LEFT JOIN stores s ON d.store_id = s.id"
    params = []
    if session['role'] != 'super_admin':
        query += " WHERE d.store_id = ?"
        params.append(session.get('store_id'))

    items = db.execute(query + " ORDER BY d.id DESC", params).fetchall()
    
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Documentos Diversos"
    cabecalho = ['Empresa', 'Número da Caixa']
    sheet.append(cabecalho)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for item in items:
        sheet.append([item['store_name'] or 'N/A', item['numero_caixa']])

    tabela = Table(displayName="TabelaDocumentos", ref=f"A1:{openpyxl.utils.get_column_letter(sheet.max_column)}{sheet.max_row}")
    estilo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tabela.tableStyleInfo = estilo
    sheet.add_table(tabela)

    for col in sheet.columns:
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = 30

    workbook.save(output)
    output.seek(0)
    return Response(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition":"attachment;filename=contas_a_pagar_documentos_diversos.xlsx"})

@app.route('/contas_a_pagar/documentos_diversos/delete/<int:item_id>', methods=['POST'])
@login_required
@role_required(['super_admin', 'admin_contas_a_pagar'])
def contas_a_pagar_diversos_delete(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM contas_a_pagar_diversos WHERE id = ?", (item_id,)).fetchone()
    if not item or (session['role'] != 'super_admin' and item['store_id'] != session.get('store_id')):
        if is_ajax_request(): return jsonify({'status': 'error', 'message': 'Documento não encontrado ou sem permissão.'}), 403
        flash('Documento não encontrado ou sem permissão.', 'danger')
    else:
        log_action('deletou_documento_diverso', target_id=item_id, target_name=f"Caixa {item['numero_caixa']}", dados_antigos=dict(item))
        db.execute("DELETE FROM contas_a_pagar_diversos WHERE id = ?", (item_id,))
        db.commit()
        if is_ajax_request(): return jsonify({'status': 'success', 'message': 'Documento excluído!', 'itemId': item_id})
        flash('Documento excluído com sucesso!', 'success')
    return redirect(url_for('documentos_diversos_dashboard'))


@app.route('/contas_a_pagar/documentos_diversos/edit/<int:item_id>', methods=['GET', 'POST'])
@login_required
@role_required(['super_admin', 'admin_contas_a_pagar'])
def contas_a_pagar_diversos_edit(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM contas_a_pagar_diversos WHERE id = ?", (item_id,)).fetchone()
    if not item or (session['role'] != 'super_admin' and item['store_id'] != session.get('store_id')):
        flash('Documento não encontrado ou sem permissão.', 'danger')
        return redirect(url_for('documentos_diversos_dashboard'))

    if request.method == 'POST':
        dados_antigos = dict(item)
        dados_novos = {'numero_caixa': request.form['numero_caixa']}
        db.execute("UPDATE contas_a_pagar_diversos SET numero_caixa = ? WHERE id = ?", (dados_novos['numero_caixa'], item_id))
        db.commit()
        log_action('editou_documento_diverso', target_id=item_id, target_name=f"Caixa {dados_novos['numero_caixa']}", dados_antigos=dados_antigos, dados_novos=dados_novos)
        flash('Documento atualizado com sucesso!', 'success')
        return redirect(url_for('documentos_diversos_dashboard'))
    return render_template('contas_a_pagar/documentos_diversos_edit.html', item=item)


if __name__ == '__main__':
    app.run(host="0.0.0.0", port=8000, debug=True)